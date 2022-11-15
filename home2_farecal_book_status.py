from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib
from tkinter import ttk,filedialog
import pandas as pd

root = Tk()
root.title('     Parking Managment System')
root.configure(bg="grey")
root.resizable(False,False)
root.geometry('975x550')

########################################################################################################################################################
def tab():
#.........................................................................................................................................................................
    def fare_calculator():
        ct1 = Canvas(root, bg="#17161b", width=975, height=470)
        ct1.place(x=1, y=1)
        equation = ""
        def show(value):
            global equation
            equation += value
            label_result.config(text=equation)
        def clear():
            global equation
            equation = ""
            label_result.config(text=equation)

        def calculate():
            vehicle = vehicle_type.get()
            day = days.get()
            if (int(vehicle) == 2):
                result = 200 * int(day)
                print(result)
                label_result.config(text=result)
            elif (int(vehicle) == 4):
                result = 600 * int(day)
                print(result)
                label_result.config(text=result)
            elif (int(vehicle) == 3):
                result = 300 * int(day)
                print(result)
                label_result.config(text=result)
            else:
                messagebox.showerror("invalid", "enter details in correct way")
        label_result = Label(root, width=25, height=2, text="", font=("arial", 30))
        label_result.place(x=164,y=20)
        Button(root, text="C", width=5, font=("arial", 30, "bold"), bd=1, fg="#fff", bg="#3697f5",
               command=lambda: clear()).place(x=220, y=140)
        Button(root, text="Calculate", width=10, font=("arial", 30, "bold"), bd=1, fg="#fff", bg="#3697f5",
               command=calculate).place(x=420, y=140)
        def on_enter(e):
            vehicle_type.delete(0, 'end')
        def on_leave(e):
            name = vehicle_type.get()
            if name == '':
                vehicle_type.insert(0, 'Enter Vehicle Type(2/3/4)')
        vehicle_type = Entry(root, width=30, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 18))
        vehicle_type.place(x=250, y=270)
        vehicle_type.insert(0, 'Enter Vehicle Type(2/3/4/other)')
        vehicle_type.bind('<FocusIn>', on_enter)
        vehicle_type.bind('<FocusOut>', on_leave)
        def on_enter(e):
            days.delete(0, 'end')
        def on_leave(e):
            name = days.get()
            if name == '':
                days.insert(0, 'Enter No of Days')
        days = Entry(root, width=30, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 18))
        days.place(x=250, y=360)
        days.insert(0, 'Enter No of Days')
        days.bind('<FocusIn>', on_enter)
        days.bind('<FocusOut>', on_leave)




#.......................................................................................................................................................................


    def booking():
        ct1 = Canvas(root, bg="#17161b", width=975, height=470)
        ct1.place(x=1, y=1)
        def clear():
            namevalue.set('')
            regvalue.set('')
            contactvalue.set('')
            emailvalue.set('')
        file = pathlib.Path('backend.xlsx')
        if file.exists():
            pass
        else:
            file = Workbook()
            sheet = file.active
            sheet['A1'] = "Full Name"
            sheet['B1'] = "Reg no"
            sheet['C1'] = "Hostel"
            sheet['D1'] = "slot"
            sheet['D1'] = "Contact number"
            sheet['E1'] = "Gender"
            sheet['F1'] = "Email id"
            file.save('backend.xlsx')
        def submit():
            name = namevalue.get()
            registration = regvalue.get()
            hostel = hostel_combobox.get()
            slot = slot_combobox.get()
            gender = gender_combobox.get()
            conatact = contactvalue.get()
            email = emailvalue.get()
            file = openpyxl.load_workbook('backend.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row + 1, value=name)
            sheet.cell(column=2, row=sheet.max_row, value=registration)
            sheet.cell(column=3, row=sheet.max_row, value=hostel)
            sheet.cell(column=4, row=sheet.max_row, value=slot)
            sheet.cell(column=5, row=sheet.max_row, value=gender)
            sheet.cell(column=6, row=sheet.max_row, value=conatact)
            sheet.cell(column=7, row=sheet.max_row, value=email)
            file.save(r'backend.xlsx')
            messagebox.showinfo('info', 'booked sucessfully')
            namevalue.set('')
            regvalue.set('')
            contactvalue.set('')
            emailvalue.set('')
        Label(root, text="Book Your Slot Now", font="arial 18", bg="#326273", fg="#fff").place(x=340, y=20)
        Label(root, text="Name", font=23, bg="#326273", fg="#fff").place(x=50, y=100)
        Label(root, text="Reg No", font=23, bg="#326273", fg="#fff").place(x=50, y=150)
        Label(root, text="Hostel", font=23, bg="#326273", fg="#fff").place(x=50, y=200)
        Label(root, text="slot no", font=23, bg="#326273", fg="#fff").place(x=430, y=200)
        Label(root, text="Gender", font=23, bg="#326273", fg="#fff").place(x=50, y=250)
        Label(root, text="Contact Number", font=23, bg="#326273", fg="#fff").place(x=50, y=300)
        Label(root, text="Email id", font=23, bg="#326273", fg="#fff").place(x=50, y=350)
        namevalue = StringVar()
        regvalue = StringVar()
        contactvalue = StringVar()
        emailvalue = StringVar()
        nameentry = Entry(root, textvariable=namevalue, width=40, bd=2, font=20)
        nameentry.place(x=200, y=100)
        regentry = Entry(root, textvariable=regvalue, width=40, bd=2, font=20)
        regentry.place(x=200, y=150)
        hostel_combobox = Combobox(root, values=['Bh1', 'Bh2', 'Gh1', 'Gh2'], font='arial 15', state='r', width=14)
        hostel_combobox.place(x=200, y=200)
        hostel_combobox.set('Bh1')
        slot_combobox = Combobox(root, values=['1', '2', '3', '4'], font='arial 15', state='r', width=14)
        slot_combobox.place(x=500, y=200)
        slot_combobox.set('1')
        contactentry = Entry(root, textvariable=contactvalue, width=40, bd=2, font=20)
        contactentry.place(x=200, y=300)
        gender_combobox = Combobox(root, values=['male', 'female'], font='arial 15', state='r', width=14)
        gender_combobox.place(x=200, y=250)
        gender_combobox.set('male')
        emailentry = Entry(root, textvariable=emailvalue, width=40, bd=2, font=20)
        emailentry.place(x=200, y=350)
        Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=200, y=400)
        Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=340, y=400)

#..............................................................................................................................................................

    def slots_available():


        ct1 = Canvas(root, bg="#17161b", width=975, height=470)
        ct1.place(x=1, y=1)

        content = Label(text="open backend.xlsx file to see booked slots",font=24)
        content.place(x=400,y=180)
        def open():
            filename = filedialog.askopenfilename(title="open a file",filetypes=(("xlxs files", ".xlsx"), ("All Files", "*.")))
            if filename:
                try:
                    filename = r"{}".format(filename)
                    df = pd.read_excel(filename)
                except:
                    messagebox.showerror("Error", "You cant access this file!")
            tree.delete(*tree.get_children())
            tree['column'] = list(df.columns)
            tree['show'] = "headings"
            for col in tree['column']:
                tree.heading(col, text=col)
            df_rows = df.to_numpy().tolist()
            for row in df_rows:
                tree.insert("", "end", values=row)
        frame = Frame(root,bg='red')
        frame.place(x=5,y=60)
        tree = ttk.Treeview(frame)
        tree.pack()
        button = Button(root, text='open', width=60, height=2, font=30, fg='white', bg="#0078d7", command=open)
        button.place(x=120,y=350)


    def logout():
       root.destroy()
       import home





    frame = Frame(root, bg='white', width=925, height=455).place(x=18, y=20)
    heading1 = Label(root,text="Welcome to Modern Parking Management System",font=('Algerian',26,'bold')).place(x=60,y=50)
    data1 = Label(root,text="In this project our team is designing a parking management system with the help of Tkinter which is a python library",font=('Algerian',9,'bold')).place(x=130,y=170)
    data2 = Label(root,text=" written in Java which is used for making GUI (graphic user interface) A parking management system automates a car ",font=('Algerian',9,'bold')).place(x=130,y=210)
    data3 = Label(root,text="parking system. It optimizes parking space and make processes efficient. It gives real-time car parking information",font=('Algerian',9,'bold')).place(x=130,y=240)
    data4 = Label(root,text=" such as vehicle & slot counts, available slots display, reserved parking, pay-and-park options, easy payments, reports,",font=('Algerian',9,'bold')).place(x=130,y=280)
    data5 = Label(root,text=" and a host of other features ",font=('Algerian',9,'bold')).place(x=130,y=320)
    Button(frame, width=15, pady=7, text="Fare Calculator", bg='white', fg='#57a1f8', border=0,font=('Microsoft YaHei UI Light', 15,), command=fare_calculator, cursor='hand2').place(x=50, y=480)
    Button(frame, width=15, pady=7, text="Book Now", bg='white', fg='#57a1f8', border=0,font=('Microsoft YaHei UI Light', 15), cursor='hand2', command=booking).place(x=320, y=480)
    Button(frame, width=15, pady=7, text="Available Slots ", bg='white', fg='#57a1f8', border=0,font=('Microsoft YaHei UI Light', 15), cursor='hand2', command=slots_available).place(x=550, y=480)
    Button(frame, width=15, pady=7, text="Log out", bg='white', fg='#57a1f8', border=0,font=('Microsoft YaHei UI Light', 15), cursor='hand2', command=logout).place(x=760, y=480)
tab()

##########################################################################################################################################################
root.mainloop()